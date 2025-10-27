import json
import re
from pathlib import Path
from openpyxl import load_workbook

# === CONFIG ===
XLSX_PATH = Path(r"./electrotehnica.xlsx")
OUTPUT_JS = XLSX_PATH.with_name("questions_from_excel.js")
SHEET_INDEX = 0          # 0 = first sheet
HEADER_ROW_SEARCH_MAX = 10  # search first N rows for headers

# Acceptable (case-insensitive) header names
Q_KEYS   = {"question", "întrebare", "intrebare"}
A_KEYS   = {"option1", "a", "var_a", "optiunea1", "opțiunea1"}
B_KEYS   = {"option2", "b", "var_b", "optiunea2", "opțiunea2"}
C_KEYS   = {"option3", "c", "var_c", "optiunea3", "opțiunea3"}
D_KEYS   = {"option4", "d", "var_d", "optiunea4", "opțiunea4"}
ANS_KEYS = {"answer", "correct", "răspuns", "raspuns", "corect"}

# Treat these RGBs as "white / not colored"
WHITE_VALUES = {"FFFFFFFF", "FFFF_FFFF", "FFffffff", "ffffff", "00000000"}

def norm(s):
    return (str(s).strip() if s is not None else "")

def lower(s):
    return norm(s).lower()

def is_solid_color_cell(cell):
    """
    Returns True if the cell looks like it has a solid non-white fill.
    Notes:
      - openpyxl stores fills via cell.fill. patternType and fgColor
      - We'll consider patternType == 'solid' and fgColor not white/empty as colored
    """
    fill = getattr(cell, "fill", None)
    if not fill:
        return False
    if getattr(fill, "patternType", None) != "solid":
        return False

    # fgColor may have .rgb, .indexed, .theme. Prefer rgb when present.
    fg = getattr(fill, "fgColor", None)
    if not fg:
        return False

    rgb = getattr(fg, "rgb", None)
    if rgb:
        rgb_up = rgb.upper().replace("-", "").replace("_", "")
        if rgb_up in WHITE_VALUES:
            return False
        # If it's exactly '00000000' (auto), ignore
        if rgb_up == "00000000":
            return False
        return True

    # If only indexed/theme: treat as colored conservatively when pattern is solid
    # (many spreadsheets with themed fills come this way)
    if getattr(fg, "indexed", None) not in (None, 64):  # 64 is default "automatic"
        return True
    if getattr(fg, "theme", None) is not None:
        return True

    return False

def find_header_row(ws):
    """Try to find a header row within the first HEADER_ROW_SEARCH_MAX rows."""
    for r in range(1, min(ws.max_row, HEADER_ROW_SEARCH_MAX) + 1):
        labels = [lower(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
        # Heuristic: row has a question key and at least 2 option keys
        has_q = any(lbl in Q_KEYS for lbl in labels)
        opt_hits = sum(lbl in (A_KEYS|B_KEYS|C_KEYS|D_KEYS) for lbl in labels)
        if has_q and opt_hits >= 2:
            return r
    # If not found, assume first row is header
    return 1

def map_columns(ws, header_row):
    """Map columns by header names. If headers are absent, assume A:question, B-E:options."""
    headers = {}
    for c in range(1, ws.max_column + 1):
        headers[c] = lower(ws.cell(row=header_row, column=c).value)

    # Try to locate columns by headers
    q_col = None
    opt_cols = []  # maintain order A,B,C,D where possible
    ans_col = None

    # Pass 1: question
    for c, lbl in headers.items():
        if lbl in Q_KEYS:
            q_col = c
            break

    # Pass 2: options (preserve an intended order A,B,C,D if labels allow)
    # Collect candidates by label groups
    label_to_col = {lbl: c for c, lbl in headers.items() if lbl}

    def pick_any(group):
        for key in group:
            if key in label_to_col:
                return label_to_col[key]
        return None

    # Try to pick A, B, C, D groups
    a_col = pick_any(A_KEYS)
    b_col = pick_any(B_KEYS)
    c_col = pick_any(C_KEYS)
    d_col = pick_any(D_KEYS)

    for col in [a_col, b_col, c_col, d_col]:
        if col and col not in opt_cols:
            opt_cols.append(col)

    # Fallback: if we didn't find headers, assume positional (A:Q, next 3-4: options)
    if not q_col and not opt_cols:
        q_col = 1  # column A
        # choose next up to 4 columns for options
        opt_cols = [2, 3, 4, 5]

    # Answer/Correct column (optional fallback)
    for c, lbl in headers.items():
        if lbl in ANS_KEYS:
            ans_col = c
            break
    return q_col, opt_cols, ans_col

def normalize_answer_text(raw, options):
    """
    Accepts:
      - exact text
      - 1..4 (human counting)
      - 0..3 (0-based)
      - A, B, C, D
    Returns text that matches one of the options, or raw if not matched.
    """
    s = norm(raw)
    if not s:
        return ""

    # numeric index?
    if re.fullmatch(r"\d+", s):
        idx = int(s)
        if 1 <= idx <= len(options):
            return options[idx - 1]
        if 0 <= idx < len(options):
            return options[idx]
    # A-D?
    if re.fullmatch(r"[A-Da-d]", s):
        idx = ord(s.upper()) - ord("A")
        if 0 <= idx < len(options):
            return options[idx]
    # otherwise treat as text
    return s

def main():
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb.worksheets[SHEET_INDEX]

    header_row = find_header_row(ws)
    q_col, opt_cols, ans_col = map_columns(ws, header_row)

    items = []
    # Iterate data rows
    for r in range(header_row + 1, ws.max_row + 1):
        q = norm(ws.cell(row=r, column=q_col).value)
        if not q:
            continue
        # Collect options (skip empties)
        opts = []
        colored_flags = []  # track which option cells are colored
        for c in opt_cols:
            cell = ws.cell(row=r, column=c)
            text = norm(cell.value)
            if text:
                opts.append(text)
                colored_flags.append(is_solid_color_cell(cell))

        # Ignore rows with fewer than 2 options
        if len(opts) < 2:
            continue

        # Determine correct answer:
        answer_text = ""
        # 1) Prefer exactly one colored option among the collected option cells
        if colored_flags and sum(colored_flags) == 1:
            idx = colored_flags.index(True)
            answer_text = opts[idx]
        else:
            # 2) Fallback to explicit Answer/Correct column if present
            if ans_col:
                raw_ans = norm(ws.cell(row=r, column=ans_col).value)
                answer_text = normalize_answer_text(raw_ans, opts)

        # Final sanity: if still empty, skip (or choose to warn)
        if not answer_text:
            # You can choose to skip or try to guess; we skip to be safe
            # print(f"Warning: no colored answer on row {r}, and no Answer column. Skipping.")
            continue
        items.append({
            "question": q,
            "options": opts,
            "answer": answer_text
        })

    if not items:
        raise SystemExit("No valid questions found. Check headers and ensure the correct option is colored (solid fill).")

    # Write JS file
    js = "const BASE_QUESTIONS = " + json.dumps(items, ensure_ascii=False, indent=2) + ";\n"
    OUTPUT_JS.write_text(js, encoding="utf-8")
    print(f"Written: {OUTPUT_JS}")

if __name__ == "__main__":
    main()
